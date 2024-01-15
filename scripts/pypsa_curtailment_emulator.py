# -*- coding: utf-8 -*-
"""
Created on 29th of November 2023
Author: Ebbe Kyhl Gøtske

This script contains the function "create_emulator" which collects
the results from the parameterization steps and updates the parameters to 
the PyPSA-Eur emulator excel file. 
Furthermore, it contains the functions used for testing and visualization of 
the emulator estimations at a user-specified scenario (currently, defined by 
the wind and solar penetration level as well as the rollout of short-duration 
and long-duration electricity storage).
"""

from openpyxl import load_workbook
from pycel import ExcelCompiler
import matplotlib.pyplot as plt
from scripts.plotting import read_and_plot_scenarios
from scripts.curtailment_parameterization import base_curtailment, technology_term, convert_series_into_2D_matrix
import numpy as np
import pandas as pd

def create_emulator(scenarios, RDIR, file_path,efficiencies,tech_labels, continuous_axis, update_excel_file=False):
    """
    This function performs the parameterization of the curtailment, first by deriving the base term of the curtailment function 
    (i.e., the curtailment as function of the primary and secondary resource) and second by deriving the technology term of the
    curtailment function (i.e., the curtailment reduction as function of the activity of a given technology).

    Inputs:
        scenarios [dict] = dictionary with the scenarios used for parameterization
        RDIR [string] = directory of the results
        file_path [string] = path to the PyPSA-Eur emulator excel file
        efficiencies [dict] = dictionary with the efficiencies of the technologies
        tech_labels [dict] = dictionary with the labels of the technologies
        update_excel_file [bool] = if True, the excel file will be updated with the new parameters

    Outputs:
        Outputs are saved to the parameters subfolder: 
            - gamma_ij_wind.csv: base term of the wind curtailment function
            - gamma_ij_solar.csv: base term of the solar curtailment function
            - beta_{tech}_wind.csv: technology term of the wind curtailment function 
            - beta_{tech}_solar.csv: technology term of the solar curtailment function 
            - wind_shares.csv: share of wind electricity of the total electricity demand
            - solar_shares.csv: share of solar PV electricity of the total electricity demand
    """
    
    variables = ["wind curt", # For parameterization step 1&2
                "solar curt",  # For parameterization step 1&2
                "wind absolute curt", # For parameterization step 1&2
                "solar absolute curt",  # For parameterization step 1&2
                "SDES dispatch", # For parameterization step 2
                "SDES absolute dispatch", # For parameterization step 2
                "SDES energy cap", # For parameterization step 2
                "SDES discharge cap", # For parameterization step 2
                "LDES dispatch", # For parameterization step 2
                "LDES absolute dispatch", # For parameterization step 2
                "LDES energy cap", # For parameterization step 2
                "LDES discharge cap", # For parameterization step 2
                "system cost", # For parameterization step 4 
                "cv wind", # For paramterization step 3
                "cv solar", # For paramterization step 3
                "backup capacity", # For paramterization step 3&4
                "backup capacity factor high", # For paramterization step 3 
                "backup capacity factor low", # For paramterization step 3
                "total demand",
                "transmission volume"
                ]

    base_case = scenarios["base"]
    ################## BASE CURTAILMENT ##############################
    print("Launching parameterization using " + continuous_axis + " axis as continuous resource...")
    df = read_and_plot_scenarios([base_case],
                                    variables_input = variables, 
                                    directory=RDIR,
                                    plot=False)

    demand = df[base_case]["total demand"].iloc[0]*1e6 # total demand in MWh - here assuming that the demand is the same for all scenarios!
    bin_lvls = [0,0.1,0.3,0.4,0.5,0.6,0.7,0.9]

    if continuous_axis == "both":
        continuous_axis1 = "primary"
        continuous_axis2 = "secondary"
        add_name1 = "_primary"
        add_name2 = "_secondary"
    else:
        continuous_axis1 = continuous_axis
        add_name1 = ""

    # wind curtailment
    var = 'wind absolute curt' # variable 
    x_name = "wind" # primary index
    gamma_ij_wind_series, x_share_df = base_curtailment(df,var,bin_lvls,demand,x_name,base_case,continuous_axis1) # base curtailment parameters
    
    # save the parameters to the parameters subfolder (for later use):
    x_share_df.to_csv("parameters/wind_shares.csv") # share of electricity
    gamma_ij_wind_series.to_csv("parameters/gamma_ij_wind" + add_name1 + ".csv") # save series to .csv

    gamma_ij_wind_2D = convert_series_into_2D_matrix(gamma_ij_wind_series,
                                                        lvls=[0,1,2,3,4,5,6],
                                                        x_i_str="wind_curtailment_w",
                                                        x_j_str="solar")
    
    gamma_ij_wind_2D.to_csv("results/gamma_ij_wind" + add_name1 + ".csv")

    if continuous_axis == "both":
        gamma_ij_wind_series2, x_share_df = base_curtailment(df,var,bin_lvls,demand,x_name,base_case,continuous_axis2) # base curtailment parameters
        gamma_ij_wind_series2.to_csv("parameters/gamma_ij_wind" + add_name2 + ".csv") # save series to .csv

        gamma_ij_wind_2_2D = convert_series_into_2D_matrix(gamma_ij_wind_series2,
                                                            lvls=[0,1,2,3,4,5,6],
                                                            x_i_str="wind_curtailment_w",
                                                            x_j_str="solar")

        gamma_ij_wind_2_2D.to_csv("results/gamma_ij_wind" + add_name2 + ".csv")

    print("Wind curtailment successfully parameterized! Proceeding to solar curtailment...")

    # solar curtailment
    var = 'solar absolute curt' # variable
    x_name = "solar" # primary index
    gamma_ij_solar_series, x_share_df = base_curtailment(df,var,bin_lvls,demand,x_name,base_case,continuous_axis1) # base curtailment parameters
    
    # save to .csv:
    x_share_df.to_csv("parameters/solar_shares.csv") # share of electricity
    gamma_ij_solar_series.to_csv("parameters/gamma_ij_solar" + add_name1 + ".csv") # save series to .csv

    if continuous_axis == "both":
        gamma_ij_solar_series2, x_share_df = base_curtailment(df,var,bin_lvls,demand,x_name,base_case,continuous_axis2) # base curtailment parameters
        gamma_ij_solar_series2.to_csv("parameters/gamma_ij_solar" + add_name2 + ".csv") # save series to .csv

    print("Solar curtailment successfully parameterized! Proceeding to technology term...")

    ########################## Technology term ##############################

    renewables = ["wind","solar"]
    techs = efficiencies.keys()
    beta_dict = {}
    for renewable in renewables:
        for tech in techs:
            scen_base = scenarios[tech][renewable][0] # base case
            scen_ref = scenarios[tech][renewable][1] # reference case without technology 
            scen_tech = scenarios[tech][renewable][2] # reference case with technology

            scens = [scen_base,
                    scen_ref,
                    scen_tech,
                    ]
            
            df_tech = read_and_plot_scenarios(scens,
                                            variables_input = variables, 
                                            directory=RDIR,
                                            plot=False)

            beta1,beta2 = technology_term(df_tech, 
                                        scen_base, scen_ref, scen_tech, 
                                        tech_name = tech, 
                                        tech_label = tech_labels[tech], 
                                        tech_efficiency = efficiencies[tech], 
                                        demand = demand,
                                        renewable=renewable)
                
            beta1.to_csv("parameters/beta_" + tech + "_" + renewable + ".csv")
            beta2.to_csv("results/beta_" + tech + "_" + renewable + ".csv")

            print(tech + " impact on " + renewable + " curtailment successfully parameterized!")

            # convert gamma_ij_wind_series and gamma_ij_solar into 2D array format:
            renewable_c = {"wind":"solar","solar":"wind"}
            beta_2d = convert_series_into_2D_matrix(beta1,
                                                    lvls=[0,1,2,3,4,5,6],
                                                    x_i_str=renewable + "_curtailment_" + renewable[0],
                                                    x_j_str=renewable_c[renewable])
            
            beta_dict[renewable, tech] = beta_2d

    print("Parameterization done!")

    ############################# Update the excel file #########################################
    # currently, the update of the Excel file can only be done for technoligies SDES and LDES.
    if update_excel_file:

        # convert gamma_ij_wind_series and gamma_ij_solar into 2D array format:
        gamma_ij_wind = convert_series_into_2D_matrix(gamma_ij_wind_series,
                                                        lvls=[0,1,2,3,4,5,6],
                                                        x_i_str="wind_curtailment_w",
                                                        x_j_str="solar")


        gamma_ij_solar = convert_series_into_2D_matrix(gamma_ij_solar_series,
                                                        lvls=[0,1,2,3,4,5,6],
                                                        x_i_str="solar_curtailment_s",
                                                        x_j_str="wind")

        beta_wind_ldes = beta_dict["wind","LDES"].copy()        
        beta_wind_ldes.index = gamma_ij_wind.index
        beta_wind_ldes.columns = gamma_ij_wind.columns

        beta_solar_ldes = beta_dict["solar","LDES"].copy()
        beta_solar_ldes.index = gamma_ij_solar.index
        beta_solar_ldes.columns = gamma_ij_solar.columns

        beta_wind_sdes = beta_dict["wind","SDES"].copy()
        beta_wind_sdes.index = gamma_ij_wind.index
        beta_wind_sdes.columns = gamma_ij_wind.columns
    
        beta_solar_sdes = beta_dict["solar","SDES"].copy()
        beta_solar_sdes.index = gamma_ij_solar.index
        beta_solar_sdes.columns = gamma_ij_solar.columns

        # Load the workbook using openpyxl
        workbook = load_workbook(file_path)

        # Select the desired sheets
        sheet_name_wind = "wind curtailment"
        sheet_name_solar = "solar curtailment"
        sheet_wind = workbook[sheet_name_wind] # sheet with wind curtailment
        sheet_solar = workbook[sheet_name_solar] # sheet with solar curtailment

        # Update the cells with the new parameter values
        for i in range(7):
            for j in range(7):
                # allocate base curtailment
                sheet_wind["D22:J28"][i][j].value = gamma_ij_wind.loc[i][j]
                sheet_solar["D22:J28"][i][j].value = gamma_ij_solar.loc[i][j]

                # allocate technology term for LDES
                sheet_wind["D51:J57"][i][j].value = beta_wind_ldes.loc[i][j]
                sheet_solar["D42:J48"][i][j].value = beta_solar_ldes.loc[i][j]

                # allocate technology term for SDES
                sheet_wind["D42:J48"][i][j].value = beta_wind_sdes.loc[i][j]
                sheet_solar["D51:J57"][i][j].value = beta_solar_sdes.loc[i][j]

        # Save the changes to the workbook
        workbook.save(file_path)

def update_excel_parameters(file_path, sheet_names, inputs):

    # Load the workbook using openpyxl
    workbook = load_workbook(file_path)

    input_values_wind = {'C13': inputs["wind"], 
                         'D13': inputs["solar"],
                         'D15': inputs["sdes"],
                         'D16': inputs["ldes"],
                         }
    
    input_values_solar = {'D13': inputs["wind"], 
                          'C13': inputs["solar"],
                          'D15': inputs["ldes"],
                          'D16': inputs["sdes"],
                          }

    # Select the desired sheets
    sheet_name_wind = sheet_names["wind"]
    sheet_name_solar = sheet_names["solar"]
    sheet_wind = workbook[sheet_name_wind] # sheet with wind curtailment
    sheet_solar = workbook[sheet_name_solar] # sheet with solar curtailment

    # Update the cell with the new parameter value
    # sheet[cell_address] = new_value
    for cell_address, new_value in input_values_wind.items():
        sheet_wind[cell_address] = new_value
        
    for cell_address, new_value in input_values_solar.items():
        sheet_solar[cell_address] = new_value

    # Save the changes to the workbook
    workbook.save(file_path)

def read_excel_outputs(file_path, output_address):

    # Load the workbook using pycel
    workbook = ExcelCompiler(file_path)

    output_addresses = ["wind curtailment" + output_address,
                        "solar curtailment" + output_address]
    # Read output values
    output = {cell_address: workbook.evaluate(cell_address) for cell_address in output_addresses}
    
    output1 = output["wind curtailment" + output_address]
    output2 = output["solar curtailment" + output_address]

    return output1, output2

def run_pypsa_emulator(file_path,curtailment_type,wind_res,solar_res,ldes,sdes):
    """ Function that updates and reads Excel sheet.
    This function is no longer used.
    """
    
    import time

    sheet_name_wind = "wind curtailment"
    sheet_name_solar = "solar curtailment"

    inputs = {'wind': wind_res,
              'solar': solar_res,
              'ldes': ldes,
              'sdes': sdes,
              }

    sheet_names = {'wind': sheet_name_wind,
                   'solar': sheet_name_solar}

    outputs_addresses = {"demand": '!L37',
                         "resources": '!L38',}
    
    output_address = outputs_addresses[curtailment_type]

    # Update input parameters
    update_excel_parameters(file_path, sheet_names, inputs)

    time.sleep(2) # currently, I need to have this much wait time to avoid data race issues
    
    # Read and print the output value
    result = read_excel_outputs(file_path, output_address)

    return result

def mask_df(df,threshold):
    df_masked = df.copy()
    df_masked[df_masked.values > threshold] = np.nan
    return df_masked

def color_plot_2D(series, renewable="", vmax=30, cbar_label="resources", disp=0.5, write_values=False):
    # convert multiindex dataframe to 2D array
    df = series.unstack()
    df = df.iloc[::-1]

    df = df.mask(df == '') # replace empty cells with NaN
    df = df.astype(float) # convert to float
    
    # plot
    cmap = "Reds"
    fig, ax = plt.subplots()
    ax.pcolormesh(df.columns,df.index,df, cmap = cmap, vmin = 0, vmax = vmax,zorder=0)

    ax.set_xlabel("wind share (% of el. demand)")
    ax.set_ylabel("solar PV share (% of el. demand)")

    # add colorbar
    cbar = plt.colorbar(ax.pcolormesh(df.columns,
                                      df.index,df, 
                                      cmap = cmap, 
                                      vmin = 0, 
                                      vmax = vmax))
    cbar.ax.set_ylabel(renewable + ' curtailment (% of ' + cbar_label + ")")

    if write_values: #len(df.index) <= 15:
        for i in range(len(df.index)):
            for j in range(len(df.columns)):
                c = df.iloc[i,j]
                if c != '' and c != np.nan and c > 0:
                    ax.text(df.columns[j]+disp, df.index[i]+disp, "{:.1f}".format(c), va='center', ha='center')

    return fig, ax, df

def plot_tech_impact(tech, renewable):
    import seaborn as sns
    beta = pd.read_csv("results/beta_" + tech + "_" + renewable + ".csv",index_col=0)
    renewable_c = {"solar":"wind",
                   "wind":"solar"}
    
    beta_array = convert_series_into_2D_matrix(beta,
                                                lvls=[0,1,2,3,4,5,6],
                                                x_i_str=renewable + "_curtailment_" + renewable[0],
                                                x_j_str=renewable_c[renewable])
    if renewable == "solar":
        beta_array = beta_array.T

    beta_array = beta_array.loc[beta_array.index[::-1]]

    fig, ax = plt.subplots(figsize=(8,6))
    im = sns.heatmap(beta_array, annot=True, ax=ax, cmap="Blues_r", cbar=True,
                     cbar_kws={'label': tech + " impact on " + renewable + " curtailment [-]"},
                     annot_kws={"size": 16},
                     fmt=".2f")

    ticklabels = [">0", "10-30", "30-40", "40-50", "50-60", "60-70", ">70"]
    ax.set_xticks(np.arange(len(ticklabels))+0.5)
    ax.set_xticklabels(ticklabels,rotation=45)
    ax.set_yticks(np.arange(len(ticklabels))+0.5)
    
    # change order of ticklabels
    ticklabels = ticklabels[::-1]
    ax.set_yticklabels(ticklabels)
    ax.set_xlabel("wind share (%)",fontsize=18)
    ax.set_ylabel("solar share (%)",fontsize=18)
                
    return fig, ax, beta_array
